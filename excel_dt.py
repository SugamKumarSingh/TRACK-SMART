from openpyxl import load_workbook
import datetime as dt

check = input("Enter your name : ")
dateTime = dt.datetime.now()
day = dateTime.day
wb = load_workbook("database.xlsx")
ws = wb.worksheets[0]

nameRange = ws.values['B2':'B27']
for cell in nameRange:
    print(cell)

    name = str(cell)
    if (name==check):
        print(name)











"""
for row in ws.iter_rows(min_row=2, max_row=26, min_col=2, max_col=2, values_only=True):
    for cell in row:
        name = str(cell)
        if (name!='None'):
            print(cell)
"""


"""
import os
import pandas as pd
from openpyxl import load_workbook
name = input("Enter your name - ")
df = pd.DataFrame({'Name' : [name]})
writer = pd.ExcelWriter('new.xlsx', engine='openpyxl')
writer.book = load_workbook('new.xlsx')
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
reader = pd.read_excel('new.xlsx')
df.to_excel(writer, index=False, header=False, startrow=len(reader)+1)
writer.close()

firebase = pyrebase.initialize_app(config)
storage = firebase.storage()

path_on_cloud = "data/demo.xlsx"
path_local="new.xlsx";
storage.child(path_on_cloud).put(path_local)
#d = os.getcwd()
#os.chdir(d)
#storage.child(path_on_cloud).download("new.xlsx")

os.remove("new.xlsx")

firebase = pyrebase.initialize_app(config)
storage = firebase.storage()

path_on_cloud = "data/demo.xlsx"
#path_local=r'D:\lol\demo.xlsx';
#storage.child(path_on_cloud).put(path_local)
#d = os.getcwd
#os.chdir(d)
storage.child(path_on_cloud).download("lol.xlsx")
"""