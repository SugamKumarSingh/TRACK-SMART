"""import pandas as pd
from openpyxl import load_workbook
writer = pd.ExcelWriter('demo.xlsx', engine='xlsxwriter')
writer.save()
name = input("Enter your name - ")
df = pd.DataFrame({'Name' : [name]})
writer = pd.ExcelWriter(r'D:\lol\demo.xlsx', engine='openpyxl')
writer.book = load_workbook('demo.xlsx')
reader = pd.read_excel(r'demo.xlsx')
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
reader = pd.read_excel(r'demo.xlsx')
df.to_excel(writer, index=False, header=False, startrow=len(reader)+1)
writer.close()
"""

import pyrebase
import os
config = {
      "apiKey": "AIzaSyAXtE0fQeJSN8r1Omtyx5vTlsdyYrF9XpE",
    "authDomain": "tympass-32736.firebaseapp.com",
    "databaseURL" : "https://tympass-32736.firebaseio.com",
    "projectId": "tympass-32736",
    "storageBucket": "tympass-32736.appspot.com",
    "messagingSenderId": "990276104410",
    "appId": "1:990276104410:web:a6d956ded09fc3c958b5e3",
    "measurementId": "G-7HF9TQ5QC1"
    }
firebase = pyrebase.initialize_app(config)
storage = firebase.storage()

path_on_cloud = "data/demo.xlsx"
#path_local=r'D:\lol\demo.xlsx';
#storage.child(path_on_cloud).put(path_local)
d = os.getcwd()
os.chdir(d)
storage.child(path_on_cloud).download("new.xlsx")

import pandas as pd
from openpyxl import load_workbook
name = input("Enter your name - ")
df = pd.DataFrame({'Name' : [name]})
writer = pd.ExcelWriter('new.xlsx', engine='openpyxl')
writer.book = load_workbook('new.xlsx')
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
reader = pd.read_excel('new.xlsx')
df.to_excel(writer, index=False, header=False, startrow=len(reader)+1)
writer.close()

firebase = pyrebase.initialize_app(config)
storage = firebase.storage()

path_on_cloud = "data/demo.xlsx"
path_local="new.xlsx";
storage.child(path_on_cloud).put(path_local)
#d = os.getcwd()
#os.chdir(d)
#storage.child(path_on_cloud).download("new.xlsx")

os.remove("new.xlsx")

firebase = pyrebase.initialize_app(config)
storage = firebase.storage()

path_on_cloud = "data/demo.xlsx"
#path_local=r'D:\lol\demo.xlsx';
#storage.child(path_on_cloud).put(path_local)
#d = os.getcwd
#os.chdir(d)
storage.child(path_on_cloud).download("lol.xlsx")
